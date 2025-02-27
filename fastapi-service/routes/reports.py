from fastapi import APIRouter, HTTPException, Response
from bson import ObjectId, Decimal128
from datetime import datetime, timedelta, timezone
from database import transactions_collection
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, PieChart

router = APIRouter()

def convert_objectid(data):
    """Convert ObjectId and Decimal128 to string or float recursively"""
    if isinstance(data, list):
        return [convert_objectid(doc) for doc in data]
    elif isinstance(data, dict):
        return {
            k: (float(v.to_decimal()) if isinstance(v, Decimal128) else str(v) if isinstance(v, ObjectId) else v)
            for k, v in data.items()
        }
    return data

async def fetch_transactions(start_date):
    """Fetch transactions from MongoDB after a given start_date."""
    transactions = await transactions_collection.find({"timestamp": {"$gte": start_date}}).to_list(1000)
    return convert_objectid(transactions)

def generate_excel_report(transactions, report_type):
    """Generate an Excel report with transaction data, line chart, and pie chart on separate sheets."""
    if not transactions:
        raise HTTPException(status_code=404, detail="No transactions found for this report.")

    df = pd.DataFrame(transactions)
    
    # Convert timestamp to datetime
    df["timestamp"] = pd.to_datetime(df["timestamp"])
    
    # Group by Date and Sum Amount
    df["date"] = df["timestamp"].dt.date
    summary = df.groupby("date")["amount"].sum().reset_index()
    
    # Group by Transaction Type
    type_summary = df.groupby("type")["amount"].sum().reset_index()

    # Create Excel Workbook
    wb = Workbook()

    # **Sheet 1: Transaction Data**
    ws_data = wb.active
    ws_data.title = "Transaction Data"

    # Write Headers
    headers = ["Transaction ID", "Sender ID", "Receiver ID", "Amount", "Type", "Timestamp"]
    ws_data.append(headers)
    
    # Write Data
    for _, row in df.iterrows():
        ws_data.append([
            row["_id"], row["senderId"], row["receiverId"], row["amount"], row["type"], row["timestamp"]
        ])

    # **Sheet 2: Line Chart (Spending Trend)**
    ws_line_chart = wb.create_sheet(title="Spending Trend")
    ws_line_chart.append(["Date", "Total Amount"])
    for _, row in summary.iterrows():
        ws_line_chart.append([row["date"], row["amount"]])

    # Create Line Chart
    line_chart = LineChart()
    line_chart.title = f"{report_type} Spending Trend"
    line_chart.y_axis.title = "Total Amount"
    line_chart.x_axis.title = "Date"
    
    data = Reference(ws_line_chart, min_col=2, min_row=1, max_row=len(summary) + 1)
    categories = Reference(ws_line_chart, min_col=1, min_row=2, max_row=len(summary) + 1)
    line_chart.add_data(data, titles_from_data=True)
    line_chart.set_categories(categories)
    
    ws_line_chart.add_chart(line_chart, "D5")

    # **Sheet 3: Pie Chart (Transaction Type Distribution)**
    ws_pie_chart = wb.create_sheet(title="Transaction Types")
    ws_pie_chart.append(["Transaction Type", "Total Amount"])
    for _, row in type_summary.iterrows():
        ws_pie_chart.append([row["type"], row["amount"]])

    # Create Pie Chart
    pie_chart = PieChart()
    pie_chart.title = "Transaction Type Distribution"
    
    type_data = Reference(ws_pie_chart, min_col=2, min_row=1, max_row=len(type_summary) + 1)
    type_labels = Reference(ws_pie_chart, min_col=1, min_row=2, max_row=len(type_summary) + 1)
    
    pie_chart.add_data(type_data, titles_from_data=True)
    pie_chart.set_categories(type_labels)
    
    ws_pie_chart.add_chart(pie_chart, "D5")  # Position the pie chart

    # Save to Bytes
    excel_bytes = BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    
    return excel_bytes

@router.get("/daily-report/excel")
async def get_daily_report_excel():
    """Generate and return a daily transaction report in Excel format."""
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    transactions = await fetch_transactions(today)
    excel_file = generate_excel_report(transactions, "Daily")
    
    return Response(
        content=excel_file.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=daily_report.xlsx"}
    )

@router.get("/weekly-report/excel")
async def get_weekly_report_excel():
    """Generate and return a weekly transaction report in Excel format."""
    last_week = datetime.now(timezone.utc) - timedelta(days=7)
    transactions = await fetch_transactions(last_week)
    excel_file = generate_excel_report(transactions, "Weekly")
    
    return Response(
        content=excel_file.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=weekly_report.xlsx"}
    )

@router.get("/monthly-report/excel")
async def get_monthly_report_excel():
    """Generate and return a monthly transaction report in Excel format."""
    last_month = datetime.now(timezone.utc) - timedelta(days=30)
    transactions = await fetch_transactions(last_month)
    excel_file = generate_excel_report(transactions, "Monthly")
    
    return Response(
        content=excel_file.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=monthly_report.xlsx"}
    )
